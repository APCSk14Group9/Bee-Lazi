from flask import Flask, request, send_file
from oct2py import octave
from sqlalchemy import and_
import os
import datetime 
import numpy
from json import *
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/home/che/test/'
global stack
stack = []
@app.route('/')
def hello():
    return """
        <!doctype html>
        <title>Upload new File</title>
        <h1>Upload new File</h1>
        <form action="client_post" method=post enctype=multipart/form-data>
          <p><input type=file name=file>
             <input type=submit value=Upload>
        </form>
        <p>%s</p>
        """ % "<br>".join(os.listdir(app.config['UPLOAD_FOLDER'], ))

@app.route('/files/<filename>')
def get_file(filename):
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename))


def newitem(s):
    t = s['condition']
    condition = ''
    for i in t:
        condition += i + '='+ '"'+t[i]+'"'+','
    condition = condition[0:-1]
    s = ( s['table']+'('+condition+')') 
    #exec(str(s))
    print(s)
    r=eval(s)
    print(r)
    return r

def select(s):
    condition =''
    t = s['condition']
    for i in t:
        condition += i +'='+ '"'+t[i]+'"'+','
    condition= condition[0:-1]  
    s = 'session.query('+s['table']+').filter_by('+condition+').all()'
    print(s)
    t = eval(s)
    return t

def update(item, s):
    for i in s:
        print('item.'+i+'='+s[i])
        exec('item.'+i+'='+s[i])

def geteventweek(s):
    print(s)
    date=datetime.date( int(s['date'][0:4]),int(s['date'][5:7]),int( s['date'][8:10]))
    #print(date)
    md = date - datetime.timedelta(date.weekday())
    sd = date + datetime.timedelta(6)
    print(md)
    return session.query(Event).filter(Event.startdate>=md).filter(Event.startdate<=sd).filter(Event.username ==s['username']).all()  

def calladaboost(s):
    import sys
    import os
    global stack
    print(stack)
    link = './Projectcs300/'
    linkfilm = link +'film/'
    linkfood = link +'food/'
    sys.path.append(os.path.abspath(linkfilm))
    sys.path.append(os.path.abspath(linkfood))
    import feedbacki, feedbacko, trainingi, trainingo
    a = session.query(Favorite).filter_by(username = s['username']).first()
    afilm = a.film()
    afood = a.food()
    afilm = [int(i) for i in afilm]
    afood = [int(i) for i in afood]
    print(afilm)
    trainingi.training(afilm)
    print('train film')
    trainingo.training(afood)
    print('train food')
    if (stack!=[]):
        feedbacko.feedback(stack)
        feedbacki.feedback(stack)
    stack = []
    import numpy as np 
    filmrank = np.load(os.path.abspath(linkfilm+'rank_list.npy'))[:13]
    print(filmrank)
    foodrank = np.load(os.path.abspath(linkfood+'rank_list.npy'))[:13]
    setitem = []
    print('ssssssss')
    import random 
    j=0
    for i in range(filmrank.__len__()):
        while(random.randint(0,1)==0 and j <foodrank.__len__()):
            setitem.append(session.query(Food).filter_by(Name = foodrank[j][0]).first())
            j=j+1
        setitem.append(session.query(Film).filter_by(Name = filmrank[i][0]).first())
        print(setitem)
    print(setitem)
    return setitem 
   
def updateitem(item,s):
    for i in s:
        print('item.'+i+'='+s[i])
        exec('item.'+i+'='+s[i])

def xuli(s):
    global stack
#    t = calladaboost(s['condition'])
    
    try:
        if (s['action'] == 'DEL'):
            setitem = select(s)
            for i in setitem:
                session.delete(i)
        if (s['action']=='INS'):
            item = newitem(s)
            session.add(item)
        if (s['action']=='SEL'):
            setitem = select(s)
            return setitem
        if (s['action']=='WEE'):
            setevent = geteventweek(s['condition'])
            return setevent
        if (s['action']=='UPD'):
            item = select(s)[0]
            updateitem(item, s['condition1'])
        if (s['action']=='XXX'):
            t = calladaboost(s['condition'])   
            return t
        if (s['action']=='LIK'):
            stack.append(1)
        if (s['action']=='UNL'):
            stack.append(0)
        session.flush()
        #session.query(User).filter_by().all()
        #exec('session.query('+s['table']+').filter_by().all()')
    except Exception as error:
        session.rollback()
        print('false'+ str(error))
        return False 
    session.commit()
    return True


@app.route('/client_post', methods=['POST'])
def add_client_file():
    #return "1"
    try:
        imgfile = request.files['file']
    except:
        imgfile = None
    if imgfile:
        filename = imgfile.filename
        imgfile.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        return str('hahaha')
    
    string = request.form['query']
    print('query:'+string)
    s = JSONDecoder().decode(string)
    res = str(xuli(s))
    print(res)
    return res

def getws(dirr):
    import openpyxl
    data1 = openpyxl.load_workbook(dirr)
    ws = data1['Sheet1']
    return ws

def loaddata():
    link = './foodlist.xlsx'
    link1 = './filmlist.xlsx'
    import numpy
    ws = getws(link)
    g= ['Type', 'Name', 'Place ', 'Address', 'Url ', 'Cost', 'Time', 'Solemnoftheplace', 'Sweet', 'Sour', 'Salty', 'Bitter', 'Spicy', 'Umami', 'Dessert', 'Privacy', 'Amount', 'Snacking', 'Amountofcalories','FoodImg']
    h =['Type', 'Name', 'Link', 'Description', 'Seriesornot', 'Duration', 'Action', 'Comedy', 'Horror', 'FamilySocialCommunity', 'Romantic', 'ScienceFiction', 'Documentary', 'Classic', 'ChildAnimation', 'Time','Linkanh'] 
    k =['foodname','place','address', 'linkanh','price']
    for i in range(2,85):
        t = Food()
        for j in range(g.__len__()):
            s = ws[chr(j+65)+str(i)].value
            cm = 't.'+str(g[j])+'="'+str(s)+'"'
            exec(cm)
        session.add(t)
    ws = getws(link1)
    sl = 88
    for i in range(2,sl):
        t = Film()
        for j in range(h.__len__()):
            s = str(ws[chr(j+65)+str(i)].value)
            s = s.replace("'","\\'")
            s = s.replace('"','\\"')
            s = s.replace('/','\\/')
#            print(s)
            cm = 't.'+str(h[j])+"='"+str(s)+"'"
            exec(cm)
        session.add(t)   
#    print(session.query(Film).filter_by().all())
    ws = getws('./addresslist.xlsx')
    for i in range(2,92):
        t = FoodAdd()
        for j in range(k.__len__()):
            s = ws[chr(j+65)+str(i)].value
            cm = 't.'+str(k[j])+'="'+str(s)+'"'
            exec(cm)
        session.add(t)
    session.commit()
 
from model import * 
if __name__ == '__main__':
    #app.run()    
    loaddata()
    app.run(host='0.0.0.0',port = 5000, debug = False)    
